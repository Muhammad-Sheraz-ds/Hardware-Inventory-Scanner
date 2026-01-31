import os
import json
import base64
import io
from datetime import datetime
from typing import List, Optional
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from groq import Groq
from openpyxl import Workbook
from dotenv import load_dotenv

load_dotenv()

app = FastAPI(title="Hardware Inventory Scanner API")

# CORS - Allow all origins for development, restrict in production
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# In-memory session storage (for demo; use Redis/DB in production)
sessions = {}

# --- Models ---
class ImageRequest(BaseModel):
    session_id: str
    image_base64: str  # Base64 encoded image data (without data:image prefix)

class HardwareData(BaseModel):
    capacity: str = "N/A"
    generation: str = "N/A"
    brand: str = "N/A"
    speed: str = "N/A"
    timestamp: Optional[str] = None

class ProcessResponse(BaseModel):
    success: bool
    data: Optional[HardwareData] = None
    error: Optional[str] = None
    scan_count: int = 0

class SessionStats(BaseModel):
    session_id: str
    scan_count: int
    started_at: str

# --- Groq Vision Processing ---
def extract_hardware_info(image_base64: str) -> dict:
    """Process image with Groq Vision API to extract hardware info"""
    api_key = os.environ.get("GROQ_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="GROQ_API_KEY not configured")
    
    client = Groq(api_key=api_key)
    
    prompt = """
    Extract the following information from this hardware label in JSON format:
    - capacity (e.g., 8GB, 16GB, 256GB)
    - generation (DDR3, DDR4, DDR5)
    - brand
    - speed (bus speed in MHz, e.g., 2133, 2400, 2666, 3200)
    If any field is missing, set it to "N/A". Return ONLY the JSON object.
    """
    
    try:
        completion = client.chat.completions.create(
            model="meta-llama/llama-4-scout-17b-16e-instruct",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {
                            "type": "image_url",
                            "image_url": {"url": f"data:image/jpeg;base64,{image_base64}"}
                        },
                    ],
                }
            ],
            response_format={"type": "json_object"}
        )
        return json.loads(completion.choices[0].message.content)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- API Endpoints ---
@app.get("/")
def health_check():
    return {"status": "healthy", "service": "Hardware Inventory Scanner API"}

@app.post("/api/start-session")
def start_session():
    """Start a new scanning session"""
    session_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    sessions[session_id] = {
        "items": [],
        "started_at": datetime.now().isoformat()
    }
    return {"session_id": session_id, "message": "Session started"}

@app.post("/api/process-image", response_model=ProcessResponse)
def process_image(request: ImageRequest):
    """Process a captured image and extract hardware information"""
    if request.session_id not in sessions:
        # Auto-create session if not exists
        sessions[request.session_id] = {
            "items": [],
            "started_at": datetime.now().isoformat()
        }
    
    try:
        # Extract hardware info from image
        result = extract_hardware_info(request.image_base64)
        
        # Add timestamp
        result["timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Store in session
        sessions[request.session_id]["items"].append(result)
        
        return ProcessResponse(
            success=True,
            data=HardwareData(**result),
            scan_count=len(sessions[request.session_id]["items"])
        )
    except HTTPException as e:
        return ProcessResponse(success=False, error=e.detail)
    except Exception as e:
        return ProcessResponse(success=False, error=str(e))

@app.get("/api/session/{session_id}")
def get_session(session_id: str):
    """Get session data"""
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    
    session = sessions[session_id]
    return {
        "session_id": session_id,
        "items": session["items"],
        "scan_count": len(session["items"]),
        "started_at": session["started_at"]
    }

@app.get("/api/export/{session_id}")
def export_session(session_id: str):
    """Export session data as Excel file"""
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    
    session = sessions[session_id]
    items = session["items"]
    
    if not items:
        raise HTTPException(status_code=400, detail="No items to export")
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Hardware Inventory"
    
    # Header row
    headers = ["#", "Brand", "Capacity", "Generation", "Speed (MHz)", "Scanned At"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = ws.cell(row=1, column=col).font.copy(bold=True)
    
    # Data rows
    for idx, item in enumerate(items, 1):
        ws.cell(row=idx + 1, column=1, value=idx)
        ws.cell(row=idx + 1, column=2, value=item.get("brand", "N/A"))
        ws.cell(row=idx + 1, column=3, value=item.get("capacity", "N/A"))
        ws.cell(row=idx + 1, column=4, value=item.get("generation", "N/A"))
        ws.cell(row=idx + 1, column=5, value=item.get("speed", "N/A"))
        ws.cell(row=idx + 1, column=6, value=item.get("timestamp", "N/A"))
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to bytes buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"hardware_inventory_{session_id}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.delete("/api/session/{session_id}")
def end_session(session_id: str):
    """End and clean up a session"""
    if session_id in sessions:
        del sessions[session_id]
    return {"message": "Session ended", "session_id": session_id}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
